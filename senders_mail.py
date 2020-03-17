#!/bin/python3

import re
from openpyxl import  Workbook
from openpyxl.utils import get_column_letter


DEFAULT_MAIL_LOG = 'maillog'
DEFAULT_TIME_PERIOD = '00:00:00-23:59:59'
DEFAULT_COUNT_MESSAGES = 1000



def check_args():
    pass


def check_params_file():
    pass


def parsing_maillog_by_id(log_data: str) -> dict: 
    '''Поиск и группировка сообщений с одинаковым ID.
    На выходе получаем словарь вида 
    {'key':['val1', 'val2',...], 
     'key':['val4', 'val5',...], 
    ...}
    key - ID сообщения
    val - полная строка из лога, принадлежащая этому ID сообщения 
    '''

    mail_processes = {}  

    for line in log_data.split('\n'):
        dirty_item_id = re.search(r'[0-9A-F]{10}:', line) #B1980961A5: 
        if dirty_item_id:
           clear_item_id = dirty_item_id.group(0)[1:] #B1980961A5 
           temp_list = []
           if clear_item_id in mail_processes.keys():
               temp_list =  mail_processes[clear_item_id]
           temp_list.append(line)
           mail_processes[clear_item_id] = temp_list
    return mail_processes


def calculate_items_by_id(mail_processes: str, reg_items: dict, search_id: str, status='sent') -> tuple: 
    ''' Функция подсчитывает количество успешных/неуспешных полученных и отправленных сообщений. 
        Подсчет ведется для доменных имен/адресатов.
        По умолчанию проверяется статус "sent".
        Входом для функции является  словарь вида:
        {'key':['val1', 'val2',...], 
         'key':['val4', 'val5',...], 
        ...}
        key - ID сообщения
        val - полная строка из лога, принадлежащая этому ID сообщения 
        Далее получаем два словаря вида
        {'dom1':num,
          'dom2':num,
          ...} 
         В первом - от кого домена/адресата  получено/оправленно, 
         во втором - для кого домена/адресата полученно/отправленно.
         На выходе получаем картеж из двух списков (FROM/TO)
         Каждый список отсортированн по количеству сообщений
         ([('examle.com', 1000), ('example2.com', 900),...], [('example3.ru', 300),...])
        '''

    from_domain = {}
    to_domain = {}
    for cur_id, cur_mess in mail_processes.items():
        t_from = ''
        t_froms = ''
        t_to_s = set()
        for mess in cur_mess:
            if not t_from:
                t_from = re.search(r'from=<[^>]+>', mess) #from=<juri@etel.ru> 
            t_to = re.search(r'to=<[^>]+>', mess) #to=<juri@etel.ru>
            status_sent = re.search(r'status={}'.format(status), mess) #
            if status_sent:
                if t_from: 
                    dom = re.search(reg_items[search_id], t_from.group(0)) #etel.ru
                    if dom: 
                        t_from_s = dom.group(0)
                if t_to: 
                    dom = re.search(reg_items[search_id], t_to.group(0)) #etel.ru
                    if dom: 
                        t_to_s.add(dom.group(0))
        if t_from_s not in from_domain.keys():
            from_domain[t_from_s] = 1
        else:
            from_domain[t_from_s] += 1 
                        
        for i_to in t_to_s:
            if i_to not in to_domain.keys():
                to_domain[i_to] = 1
            else:
                to_domain[i_to] += 1 

    sorted_from_domain = sorted(from_domain.items(), key=lambda kv: kv[1], reverse=True)
    sorted_to_domain = sorted(to_domain.items(), key=lambda kv: kv[1], reverse=True)
    return sorted_from_domain, sorted_to_domain


def save_report_loads_mails_domains_to_excel(*data_reports: list) -> None:
    '''Принимает на вход список доменов/адресатов вида (domain, num) тип tuple
    первый элемент - заголовок столбца(напр."FROM DOMAIN").
    сохраняет данный отчет в листе excel.'''

    file_name = 'report_result_load_mail_domains.xlsx'
    wb = Workbook()
    work_sheet = wb.active
    work_sheet.title = 'report 1'
    col = 0
    for data_report in data_reports:
        col += 1
        for item_domains in data_report:
            for row, domain in enumerate(item_domains, 1):
                work_sheet.cell(column=col, row=row, value=domain[0])
                work_sheet.cell(column=col+1, row=row, value=domain[1])
            col += 2
    wb.save(filename=file_name)


def main():
    
    reg_items = {
            'domain': '@([a-zA-Z0-9-_]+\.+)*[a-z]{2,6}',
            'address': '[a-zA-Z0-9-_=+\.]+@([a-zA-Z0-9-_]+\.+)*[a-z]{2,6}',
            }

    m_log_file = DEFAULT_MAIL_LOG

    with open(m_log_file, encoding='utf-8', errors='ignore') as log_file:
        log_data = log_file.read()
    
    messages_by_id = parsing_maillog_by_id(log_data)

    (sorted_from_domain, sorted_to_domain) = calculate_items_by_id(messages_by_id, reg_items,  search_id='domain')
    (sorted_from_addr, sorted_to_addr) = calculate_items_by_id(messages_by_id, reg_items,  search_id='address')
        
    temp_from_domains = [('FROM DOMAIN', 'NUM')]
    temp_to_domains = [('TO DOMAIN', 'NUM')]
    temp_from_domains.extend(sorted_from_domain)
    temp_to_domains.extend(sorted_to_domain)

    temp_from_addr= [('FROM ADDR', 'NUM')]
    temp_to_addr= [('TO ADDR', 'NUM')]
    temp_from_addr.extend(sorted_from_addr)
    temp_to_addr.extend(sorted_to_addr)
    
    save_report_loads_mails_domains_to_excel((temp_from_domains, temp_to_domains), (temp_from_addr, temp_to_addr))


if __name__ == "__main__":
    main()
